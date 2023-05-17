import json
import pickle
# ---
# Libraries Used For Animation
import time
from pathlib import Path

# ---
import emoji  # Emoji detecter
import matplotlib.pyplot as plt
import pandas as pd
# ---
# Libraries Used For Data Visualization
import plotly.express as px
import plotly.graph_objects as go
import requests
import seaborn as sns
import streamlit as st
# ---
# Libraries Used For Autentication
import streamlit_authenticator as stauth
from streamlit_lottie import st_lottie
# ---
from streamlit_option_menu import option_menu
import openpyxl
from datetime import datetime
from datetime import timedelta
#wb=openpyxl.load_workbook()
#--------------

#trigger
import schedule
import time
import datetime

st.set_page_config(page_title="Dashboard")

st.sidebar.title(f"Hoşgeldin Yeni Nesil Teknolojiler")
with st.sidebar:
    page = option_menu("Menü", ["Yeni Nesil Genel", "SLA Analiz", "Elif Ceyhan", "Sevtap Sevgili", "Ece Aksen", "Yağmur Erışık",  "Adnan Çomakoğlu"  ],
                       icons=['house','gear'], menu_icon="cast", default_index=0)

def job():
    print("Zaman:", datetime.datetime.now())


#schedule.every().monday.at("10:00").do(job)
schedule.every(1).minutes.do(job)

#triggerend
while True:
    schedule.run_pending()
    time.sleep(1)

    #Data Pereparation
    #df=pd.read_excel('/Users/eceaks/PycharmProjects/streamlite/Rapor.xlsx')

    wb=openpyxl.load_workbook('/Users/eceaks/PycharmProjects/streamlite/Rapor_O.xlsx')
    #wb=openpyxl.load_workbook('//ZHLEVRPADW022/RPA_Projeler/BTYP_Dashboard/Rapor_O.xlsx')

    ws=wb["Rapor_O"]
    tarihbilgisi=ws["B2"].value
    ws["Q9"]=tarihbilgisi
    ws["Q9"]=datetime.datetime.now()
    ws.delete_rows(1,8)
    #-
    ws=wb.active
    constant_cell=ws["Q1"]
    fixed_value=constant_cell.value
    ws["O1"]="SLA"

    for row in range(2, ws.max_row+1):
        cell_value=ws.cell(row=row,column=8).value
        difference=cell_value-fixed_value
        absolute_diff = abs(difference.total_seconds() / 86400)
        ws.cell(row=row, column=15).value=absolute_diff

    ws.cell(row=1,column=1).value='Talep_No'
    ws.cell(row=1,column=2).value='Talep_Sınıflandırma'
    ws.cell(row=1,column=3).value='Talep_Açıklaması'
    ws.cell(row=1,column=4).value='Sektör'
    ws.cell(row=1,column=5).value='Sirket'
    ws.cell(row=1,column=6).value='Talep_Sahibi	'
    ws.cell(row=1,column=7).value='Talep_Durumu'
    ws.cell(row=1,column=8).value='Talep_Baslangıc_Tarihi'
    ws.cell(row=1,column=9).value='Talep_Kapanıs_Tarihi'
    ws.cell(row=1,column=10).value='Kaynak_Yöneticisi'
    ws.cell(row=1,column=11).value='Toplam_Aktivite_Süresi'
    ws.cell(row=1,column=12).value='Talep_Tipi'
    ws.cell(row=1,column=13).value='Talep_Sınıfı'
    ws.cell(row=1,column=14).value='Talep_Modülü'

    ws["P1"]="Talep_Statü"
    for row in range(2, ws.max_row+1):
        cell=ws.cell(row=row,column=7)
        if cell.value=="Sonlandırıldı":
            ws.cell(row=row,column=16).value="Kapalı"
        else:
            ws.cell(row=row,column=16).value="Açık"

    wb.save('/Users/eceaks/PycharmProjects/streamlite/RaporEdited.xlsx')


    #pg config eski konum

    dataset=st.container()
    with dataset:
        datasource = '/Users/eceaks/PycharmProjects/streamlite/RaporEdited.xlsx'
        Sheet1 = 'Rapor_O'
        df1 = pd.read_excel(datasource, sheet_name=Sheet1, usecols='A:P', header=0)
        #Data Prep
        Talep_Sınıfı_Listesi = df1['Talep_Sınıfı'].unique().tolist()
        Talep_Modülü_Listesi = df1['Talep_Modülü'].unique().tolist()
        Kapalı_Talep_df1 = df1[df1["Talep_Statü"] == 'Kapalı']
        Acık_Talep_df1 = df1[df1.Talep_Statü == 'Açık']
        Acık_Talep_Sayısı = Acık_Talep_df1["Talep_Statü"].value_counts()
        Kapalı_Talep_Sayısı= Kapalı_Talep_df1["Talep_Statü"].value_counts()
        Tüm_Taleplerin_Sayısı_df1 = df1["Talep_Statü"].value_counts()
        #df2 = pd.DataFrame()
        #df2 = pd.concat([Acık_Talep_Sayısı, Kapalı_Talep_Sayısı], axis=0)
        #df2 = pd.DataFrame(df2, columns=["Talep_Durumu"])



    #side bar eski konum


    if page=="Yeni Nesil Genel":
        st.markdown("<h1 style='text-align:center;'>YENİ NESİL TEKNOLOJİLER</h1>", unsafe_allow_html=True)
        st.title("İstek Performans Raporu")
        st.caption("Analizler BTYP Raporları kullanılarak oluşturulmuştur.")
        st.caption("Projeye Bağlanmamış Uygulama Destek Taleplerinin Detay Raporudur.")
        st.subheader("Talep Listesi")
        st.write(df1)
        st.subheader("Kapalı Talepler")
        st.write(Kapalı_Talep_df1)
        st.subheader("Açık Talepler")
        st.write( Acık_Talep_df1)
        st.subheader("Açık Talep Sayısı")
        st.write( Acık_Talep_Sayısı)
        st.subheader("Kapalı Talep Sayısı")
        st.write( Kapalı_Talep_Sayısı)

        #------------------------------------------------------------
        st.subheader("Talep Dağılımı")
        fig = go.Figure(data=[go.Pie(labels=['Açık', 'Kapalı'], values=Tüm_Taleplerin_Sayısı_df1, hole=.3)])
        fig.update_layout(title_text="Açık vs Kapalı")
        st.plotly_chart(fig)

        st.subheader('Talep Sınıfına Göre Dağılım')
        # Inserted expander
        with st.expander("Görmek için Tıkla -> Her Sınıf Tipi İçin Talep Sayısı"):
            st.write(df1["Talep_Sınıfı"].value_counts())

        st.subheader('Talep Modülüne Göre Dağılım')
        with st.expander("Görmek için Tıkla -> Her Modül Tipi İçin Talep Sayısı"):
            st.write(df1["Talep_Modülü"].value_counts())

        # ---#
        tabv1, tabv2 = st.tabs(['Talep_Sınıfı', "Talep_Modülü"])
        with tabv1:
            # ---------
            st.subheader("Her Sınıf Tipinden Gelen Talep Sayısı")
            a = df1['Talep_Sınıfı'].value_counts().reset_index()
            a = pd.DataFrame(a)
            a.columns = ['Talep_Sınıfı', 'NumberofRequest']
            pie_chart2 = px.pie(a, title='Talep Dağılımı', values=a['NumberofRequest'], names='Talep_Sınıfı')
            st.plotly_chart(pie_chart2)
        with tabv2:
            st.header("Her Modül Tipinden Gelen İstek Sayısı")
            b = df1['Talep_Modülü'].value_counts().reset_index()
            b = pd.DataFrame(b)
            b.columns = ['Talep_Modülü', 'NumberofRequest']
            pie_chart3 = px.pie(b, title='Her Modül için Talep Dağılımı', values=b['NumberofRequest'],
                                names='Talep_Modülü')
            st.plotly_chart(pie_chart3)
            # ---#

        st.subheader("Modül Tiplerine göre Kapalı & Açık Taleplerin Dağılımı")
        tab1, tab2 = st.tabs(["Kapalı Talepler", "Açık Talepler"])
        with tab1:
            st.write("Modül Tipi / Kapalı Talep Sayısı")
            Talep_Modülü_nmbr_cr = pd.DataFrame(Kapalı_Talep_df1['Talep_Modülü'].value_counts())
            st.bar_chart(Talep_Modülü_nmbr_cr)
        with tab2:
            st.write("Modül Tipi / Açık Talep Sayısı")
            Talep_Modülü_nmbr_or = pd.DataFrame(Acık_Talep_df1['Talep_Modülü'].value_counts())
            st.bar_chart(Talep_Modülü_nmbr_or)



        kaynak_yöneticisi = pd.DataFrame(Acık_Talep_df1['Kaynak_Yöneticisi'].value_counts())
        st.bar_chart(kaynak_yöneticisi)
        with st.expander("Detaylı İncelemek için"):
            st.write(kaynak_yöneticisi)


        # Detailed Pie CHart Anaylsis
        st.subheader("Detay Analizler")
        path = st.multiselect("Değişken Seçimi:", (
            'Talep_Durumu', 'Talep_Sahibi','Talep_Sınıflandırma' , 'Sirket', 'Talep_Tipi','Talep_Sınıfı',
            'Kaynak_Yöneticisi', 'Talep_Modülü'))
        fig = px.sunburst(data_frame=df1, path=path)
        st.plotly_chart(fig)
        # ---#





    if page=="Elif Ceyhan":

        st.header("Açık Talepler")
        st.caption("Tüm Açık Talepler")
        Member_acık = Acık_Talep_df1[Acık_Talep_df1['Kaynak_Yöneticisi'].str.contains("Elif")]
        st.write(Member_acık)

        # ---Sla süresi açık taleplerin
        st.caption("Sla Süresi 5 Günü Geçen Açık Talepler")
        Member_acık_SLA=Member_acık[Member_acık["SLA"] >30]
        st.write(Member_acık_SLA)
        st.caption("Sla Süresi 10 Günü Geçen Açık Talepler")
        Member_acık_SLA = Member_acık[Member_acık["SLA"] > 30]
        st.write(Member_acık_SLA)
        st.caption("Sla Süresi 30 Günü Geçen Açık Talepler")
        Member_acık_SLA = Member_acık[Member_acık["SLA"] > 30]
        st.write(Member_acık_SLA)

        st.header("Kapalı Talepler")
        st.caption("Tüm Kapalı Talepler")
        Member_kapalı = Kapalı_Talep_df1[Kapalı_Talep_df1['Kaynak_Yöneticisi'].str.contains("Elif")]
        st.write(Member_kapalı)

        #---Sla süresi kapalı taleplerin
        st.caption("Sla Süresi 5 Günü Geçen Kapalı Talepler")
        Member_kapalı_SLA = Member_kapalı[Member_kapalı["SLA"] > 5]
        st.write(Member_kapalı_SLA)
        st.caption("Sla Süresi 10 Günü Geçen Kapalı Talepler")
        Member_kapalı_SLA = Member_kapalı[Member_kapalı["SLA"] > 10]
        st.write(Member_kapalı_SLA)
        st.caption("Sla Süresi 30 Günü Geçen Kapalı Talepler")
        Member_kapalı_SLA = Member_kapalı[Member_kapalı["SLA"] > 30]
        st.write(Member_kapalı_SLA)

        st.subheader("Taleplerin İletilme Tarihleri")
        df1["Talep_Baslangıc_Tarihi"] = pd.to_datetime(df1["Talep_Baslangıc_Tarihi"])
        df_grouped_count = df1.groupby(df1["Talep_Baslangıc_Tarihi"].dt.date).Talep_No.count()
        st.bar_chart(df_grouped_count)
        st.write(df_grouped_count)
        df_grouped = df1.groupby("Talep_Baslangıc_Tarihi")["Talep_No"]



    if page=="Ece Aksen":
        st.header("Açık Talepler")
        st.caption("Tüm Açık Talepler")
        Member_acık = Acık_Talep_df1[Acık_Talep_df1['Kaynak_Yöneticisi'].str.contains("Ece")]
        st.write(Member_acık)

        # ---Sla süresi açık taleplerin
        st.caption("Sla Süresi 5 Günü Geçen Açık Talepler")
        Member_acık_SLA = Member_acık[Member_acık["SLA"] > 30]
        st.write(Member_acık_SLA)
        st.caption("Sla Süresi 10 Günü Geçen Açık Talepler")
        Member_acık_SLA = Member_acık[Member_acık["SLA"] > 30]
        st.write(Member_acık_SLA)
        st.caption("Sla Süresi 30 Günü Geçen Açık Talepler")
        Member_acık_SLA = Member_acık[Member_acık["SLA"] > 30]
        st.write(Member_acık_SLA)

        st.header("Kapalı Talepler")
        st.caption("Tüm Kapalı Talepler")
        Member_kapalı = Kapalı_Talep_df1[Kapalı_Talep_df1['Kaynak_Yöneticisi'].str.contains("Ece")]
        st.write(Member_kapalı)

        # ---Sla süresi kapalı taleplerin
        st.caption("Sla Süresi 5 Günü Geçen Kapalı Talepler")
        Member_kapalı_SLA = Member_kapalı[Member_kapalı["SLA"] > 5]
        st.write(Member_kapalı_SLA)
        st.caption("Sla Süresi 10 Günü Geçen Kapalı Talepler")
        Member_kapalı_SLA = Member_kapalı[Member_kapalı["SLA"] > 10]
        st.write(Member_kapalı_SLA)
        st.caption("Sla Süresi 30 Günü Geçen Kapalı Talepler")
        Member_kapalı_SLA = Member_kapalı[Member_kapalı["SLA"] > 30]
        st.write(Member_kapalı_SLA)

        st.subheader("Taleplerin İletilme Tarihleri")
        df1["Talep_Baslangıc_Tarihi"] = pd.to_datetime(df1["Talep_Baslangıc_Tarihi"])
        df_grouped_count = df1.groupby(df1["Talep_Baslangıc_Tarihi"].dt.date).Talep_No.count()
        st.bar_chart(df_grouped_count)
        st.write(df_grouped_count)
        df_grouped = df1.groupby("Talep_Baslangıc_Tarihi")["Talep_No"]

    if page=="Sevtap Sevgili":
        st.header("Açık Talepler")
        st.caption("Tüm Açık Talepler")
        Member_acık = Acık_Talep_df1[Acık_Talep_df1['Kaynak_Yöneticisi'].str.contains("Sevtap")]
        st.write(Member_acık)

        # ---Sla süresi açık taleplerin
        st.caption("Sla Süresi 5 Günü Geçen Açık Talepler")
        Member_acık_SLA = Member_acık[Member_acık["SLA"] > 30]
        st.write(Member_acık_SLA)
        st.caption("Sla Süresi 10 Günü Geçen Açık Talepler")
        Member_acık_SLA = Member_acık[Member_acık["SLA"] > 30]
        st.write(Member_acık_SLA)
        st.caption("Sla Süresi 30 Günü Geçen Açık Talepler")
        Member_acık_SLA = Member_acık[Member_acık["SLA"] > 30]
        st.write(Member_acık_SLA)

        st.header("Kapalı Talepler")
        st.caption("Tüm Kapalı Talepler")
        Member_kapalı = Kapalı_Talep_df1[Kapalı_Talep_df1['Kaynak_Yöneticisi'].str.contains("Sevtap")]
        st.write(Member_kapalı)

        # ---Sla süresi kapalı taleplerin
        st.caption("Sla Süresi 5 Günü Geçen Kapalı Talepler")
        Member_kapalı_SLA = Member_kapalı[Member_kapalı["SLA"] > 5]
        st.write(Member_kapalı_SLA)
        st.caption("Sla Süresi 10 Günü Geçen Kapalı Talepler")
        Member_kapalı_SLA = Member_kapalı[Member_kapalı["SLA"] > 10]
        st.write(Member_kapalı_SLA)
        st.caption("Sla Süresi 30 Günü Geçen Kapalı Talepler")
        Member_kapalı_SLA = Member_kapalı[Member_kapalı["SLA"] > 30]
        st.write(Member_kapalı_SLA)

        st.subheader("Taleplerin İletilme Tarihleri")
        df1["Talep_Baslangıc_Tarihi"] = pd.to_datetime(df1["Talep_Baslangıc_Tarihi"])
        df_grouped_count = df1.groupby(df1["Talep_Baslangıc_Tarihi"].dt.date).Talep_No.count()
        st.bar_chart(df_grouped_count)
        st.write(df_grouped_count)
        df_grouped = df1.groupby("Talep_Baslangıc_Tarihi")["Talep_No"]

    if page=="Yağmur Erışık":
        st.header("Açık Talepler")
        st.caption("Tüm Açık Talepler")
        Member_acık = Acık_Talep_df1[Acık_Talep_df1['Kaynak_Yöneticisi'].str.contains("Yağmur")]
        st.write(Member_acık)

        # ---Sla süresi açık taleplerin
        st.caption("Sla Süresi 5 Günü Geçen Açık Talepler")
        Member_acık_SLA = Member_acık[Member_acık["SLA"] > 30]
        st.write(Member_acık_SLA)
        st.caption("Sla Süresi 10 Günü Geçen Açık Talepler")
        Member_acık_SLA = Member_acık[Member_acık["SLA"] > 30]
        st.write(Member_acık_SLA)
        st.caption("Sla Süresi 30 Günü Geçen Açık Talepler")
        Member_acık_SLA = Member_acık[Member_acık["SLA"] > 30]
        st.write(Member_acık_SLA)

        st.header("Kapalı Talepler")
        st.caption("Tüm Kapalı Talepler")
        Member_kapalı = Kapalı_Talep_df1[Kapalı_Talep_df1['Kaynak_Yöneticisi'].str.contains("Yağmur")]
        st.write(Member_kapalı)

        # ---Sla süresi kapalı taleplerin
        st.caption("Sla Süresi 5 Günü Geçen Kapalı Talepler")
        Member_kapalı_SLA = Member_kapalı[Member_kapalı["SLA"] > 5]
        st.write(Member_kapalı_SLA)
        st.caption("Sla Süresi 10 Günü Geçen Kapalı Talepler")
        Member_kapalı_SLA = Member_kapalı[Member_kapalı["SLA"] > 10]
        st.write(Member_kapalı_SLA)
        st.caption("Sla Süresi 30 Günü Geçen Kapalı Talepler")
        Member_kapalı_SLA = Member_kapalı[Member_kapalı["SLA"] > 30]
        st.write(Member_kapalı_SLA)

        st.subheader("Taleplerin İletilme Tarihleri")
        df1["Talep_Baslangıc_Tarihi"] = pd.to_datetime(df1["Talep_Baslangıc_Tarihi"])
        df_grouped_count = df1.groupby(df1["Talep_Baslangıc_Tarihi"].dt.date).Talep_No.count()
        st.bar_chart(df_grouped_count)
        st.write(df_grouped_count)
        df_grouped = df1.groupby("Talep_Baslangıc_Tarihi")["Talep_No"]

    if page=="Adnan Çomakoğlu":
        st.header("Açık Talepler")
        st.caption("Tüm Açık Talepler")
        Member_acık = Acık_Talep_df1[Acık_Talep_df1['Kaynak_Yöneticisi'].str.contains("Adnan")]
        st.write(Member_acık)

        # ---Sla süresi açık taleplerin
        st.caption("Sla Süresi 5 Günü Geçen Açık Talepler")
        Member_acık_SLA = Member_acık[Member_acık["SLA"] > 30]
        st.write(Member_acık_SLA)
        st.caption("Sla Süresi 10 Günü Geçen Açık Talepler")
        Member_acık_SLA = Member_acık[Member_acık["SLA"] > 30]
        st.write(Member_acık_SLA)
        st.caption("Sla Süresi 30 Günü Geçen Açık Talepler")
        Member_acık_SLA = Member_acık[Member_acık["SLA"] > 30]
        st.write(Member_acık_SLA)

        st.header("Kapalı Talepler")
        st.caption("Tüm Kapalı Talepler")
        Member_kapalı = Kapalı_Talep_df1[Kapalı_Talep_df1['Kaynak_Yöneticisi'].str.contains("Adnan")]
        st.write(Member_kapalı)

        # ---Sla süresi kapalı taleplerin
        st.caption("Sla Süresi 5 Günü Geçen Kapalı Talepler")
        Member_kapalı_SLA = Member_kapalı[Member_kapalı["SLA"] > 5]
        st.write(Member_kapalı_SLA)
        st.caption("Sla Süresi 10 Günü Geçen Kapalı Talepler")
        Member_kapalı_SLA = Member_kapalı[Member_kapalı["SLA"] > 10]
        st.write(Member_kapalı_SLA)
        st.caption("Sla Süresi 30 Günü Geçen Kapalı Talepler")
        Member_kapalı_SLA = Member_kapalı[Member_kapalı["SLA"] > 30]
        st.write(Member_kapalı_SLA)

        st.subheader("Taleplerin İletilme Tarihleri")
        df1["Talep_Baslangıc_Tarihi"] = pd.to_datetime(df1["Talep_Baslangıc_Tarihi"])
        df_grouped_count = df1.groupby(df1["Talep_Baslangıc_Tarihi"].dt.date).Talep_No.count()
        st.bar_chart(df_grouped_count)
        st.write(df_grouped_count)
        df_grouped = df1.groupby("Talep_Baslangıc_Tarihi")["Talep_No"]

    if page=="SLA Analiz":
        st.title("SLA Süre Analizleri")
        # ----------------------------------------------------
        st.subheader("Çözüm Süresi 5 günü geçen talepler")
        print(Acık_Talep_df1)
        df_Sla5 = Acık_Talep_df1[Acık_Talep_df1["SLA"] > 5]
        st.write(df_Sla5)
        df_Sla5_count = df_Sla5["Talep_Sınıfı"].value_counts()
        df_Sla5_count1 = df_Sla5["Talep_Modülü"].value_counts()
        df_Sla5_count2 = df_Sla5["Talep_Statü"].value_counts()
        st.write(df_Sla5_count)
        st.write(df_Sla5_count1)
        st.write(df_Sla5_count2)

        # ----------------------------------------------------
        st.subheader("Çözüm Süresi 30 günü geçen talepler")
        print(Acık_Talep_df1)
        df_Sla30 = Acık_Talep_df1[Acık_Talep_df1["SLA"] > 30]
        st.write(df_Sla30)
        df_Sla30_count = df_Sla30["Talep_Sınıfı"].value_counts()
        df_Sla30_count1 = df_Sla5["Talep_Modülü"].value_counts()
        df_Sla30_count2 = df_Sla30["Talep_Statü"].value_counts()
        st.write(df_Sla30_count)
        st.write(df_Sla30_count1)
        st.write(df_Sla30_count2)

        # ----------------------------------------------------
        st.subheader("Çözüm Süresi 50 günü geçen talepler")
        print(Acık_Talep_df1)
        df_Sla50 = Acık_Talep_df1[Acık_Talep_df1["SLA"] > 50]
        st.write(df_Sla50)
        df_Sla50_count = df_Sla50["Talep_Sınıfı"].value_counts()
        df_Sla50_count1 = df_Sla5["Talep_Modülü"].value_counts()
        df_Sla50_count2 = df_Sla50["Talep_Statü"].value_counts()
        st.write(df_Sla50_count)
        st.write(df_Sla50_count1)
        st.write(df_Sla50_count2)

        # --------------
        st.header("Ay Bazında Gelen Talep Sayısı")
        st.write(emoji.emojize(':chart_increasing:'))
        line = pd.DataFrame(df1['Talep_Baslangıc_Tarihi'].value_counts())
        st.bar_chart(line)
        with st.expander("Detaylı İncelemek için"):
            st.write(line)

        st.header("Ay Bazında Gelen Kapatılan Talep Sayısı")
        st.write(emoji.emojize(':chart_increasing:'))
        line = pd.DataFrame(Kapalı_Talep_df1['Talep_Kapanıs_Tarihi'].value_counts())
        st.bar_chart(line)
        with st.expander("Detaylı İncelemek için"):
            st.write(line)

        # --------------

        st.subheader("Sınıf Tiplerinin Ortalama Sla Süresi")
        grouped = df1.groupby("Talep_Sınıfı")["SLA"]
        mean_Sla = grouped.mean()
        st.write( mean_Sla)
        st.subheader("Modül Tipleri Ortalama Sla Süresi")
        grouped = df1.groupby("Talep_Modülü")["SLA"]
        mean_Sla = grouped.mean()
        st.write( mean_Sla)
        #-----------------------------------
        # Scatter Plot
        st.subheader("Sla Süresi / Talep Numarası")
        st.write("Opsiyonlar: Sınıf Tipi, Modül Tipi, Talep Tipi")

        # Version 1 Scatter Plot
        selectopt = st.selectbox('Kategori Seçimi', ("Talep_Sınıfı", "Modül_Tipi", "İstek_Tipi_Detay"))
        fig = px.scatter(data_frame=df1, x='Talep_No', y="SLA", color=selectopt)
        st.plotly_chart(fig)

        # Summary Table
        st.subheader('Talep Tarihi/SLA/Talep Sınıfı Detay Rapor')
        fig = go.Figure(data=go.Table(
            header=dict(values=list(df1[['Talep_Baslangıc_Tarihi', "Talep_No", 'SLA', 'Talep_Sınıfı']].columns),
                        fill_color='#FD8E72', align='center'),
            cells=dict(values=[df1.Talep_Baslangıc_Tarihi, df1.Talep_No, df1.SLA, df1.Talep_Sınıfı],
                       fill_color='#E5ECF6',
                       align='left')))
        # fig.update_layout()
        st.write(fig)
        # -----------------------------------------


        st.subheader("Her Sınıf Tipi için SLA Süre Analizi ")
        tab1, tab2 = st.tabs(["Versiyon 1", "Versiyon 2"])
        #kontrol
        # Box Plot Version 1
        chart2 = ('RPA', 'ZES EV IOT', 'MOBİL UYGULAMALAR', 'INTRANET SİTELERİ', 'Optimizasyon Portali')
        chart_selection2 = st.selectbox("Sınıf Tipi Seçin1", chart2)
        dataf2rpa = df1[df1["Talep_Sınıfı"] == chart_selection2]
        fig = px.box(data_frame=dataf2rpa, x="Talep_Sınıfı", y="SLA", color="Talep_Sınıfı")
        st.plotly_chart(fig)

        st.subheader("Talep vs SLA Süresi")
        unique_SLA = df1['SLA'].unique().tolist()
        moduleselect = st.slider('Sla Süresi', min_value=min(unique_SLA), max_value=max(unique_SLA),
                                 value=(min(unique_SLA), max(unique_SLA)))

        typeofmodul = st.multiselect('Birden Fazla Sınıf Tipi Seçebilirsiniz', Talep_Sınıfı_Listesi,
                                     default=Talep_Sınıfı_Listesi)
        # !tryout = (df1["SLA"].between(*moduleselect)) & (df1['Talep_Sınıfı'].isin(typeofmodul))
        # !number_of_result = df1[tryout].shape[0]
        # !st.markdown(f'*Belirlenmiş Sla Süresi Aralığında Bulunan Mevcut Talep sayısı : {number_of_result}*')



